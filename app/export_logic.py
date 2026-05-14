from __future__ import annotations

import re
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload

from app.models import Report


def _safe_decimal(value: object) -> float:
    if value is None:
        return 0.0
    if isinstance(value, Decimal):
        return float(value)
    return float(value)


def _safe_text(value: object) -> str:
    if value is None:
        return ""
    return str(value)


def _safe_filename(value: str) -> str:
    cleaned = re.sub(r'[\\/*?:"<>|]+', "_", value).strip()
    cleaned = re.sub(r"\s+", " ", cleaned)
    return cleaned or "report"


def export_report_xlsx(db: Session, report_id: int) -> tuple[BytesIO, str]:
    report = db.execute(
        select(Report)
        .options(selectinload(Report.lines))
        .where(Report.id == report_id)
    ).scalar_one_or_none()

    if report is None:
        raise ValueError("Report not found")

    wb = Workbook()
    ws = wb.active
    ws.title = "Calculated_Data_Report"

    headers = [
        "Item Name",
        "SKU",
        "Manufacturer",
        "Vendor Code",
        "Category",
        "Rate",
        "Stock_Available",
        "Quantity",
        "Build Cost",
        "Qty TBO",
        "Total Cost",
        "Related Composite",
    ]

    ws.append(headers)

    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    wrap_alignment = Alignment(wrap_text=True, vertical="top")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = wrap_alignment

    first_data_row = 2
    for row_offset, line in enumerate(report.lines):
        excel_row = first_data_row + row_offset
        ws.append(
            [
                _safe_text(line.item_name),
                _safe_text(line.sku),
                _safe_text(line.manufacturer),
                _safe_text(line.vendor_code),
                _safe_text(line.category_name),
                _safe_decimal(line.rate),
                _safe_decimal(line.stock_available),
                _safe_decimal(line.quantity),
                f"=F{excel_row}*H{excel_row}",
                _safe_decimal(line.qty_tbo),
                _safe_decimal(line.total_cost),
                _safe_text(line.related_composite),
            ]
        )

    last_data_row = ws.max_row
    has_data_rows = last_data_row >= first_data_row

    if has_data_rows:
        totals_row = last_data_row + 1
        ws.cell(row=totals_row, column=1, value="Итого")
        ws.cell(
            row=totals_row,
            column=9,
            value=f"=SUM(I{first_data_row}:I{last_data_row})",
        )
        ws.cell(
            row=totals_row,
            column=11,
            value=f"=SUM(K{first_data_row}:K{last_data_row})",
        )

        totals_font = Font(bold=True)
        totals_fill = PatternFill(fill_type="solid", fgColor="F2F2F2")
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=totals_row, column=col_idx)
            cell.font = totals_font
            cell.fill = totals_fill

    widths = {
        "A": 45,
        "B": 18,
        "C": 28,
        "D": 18,
        "E": 28,
        "F": 12,
        "G": 16,
        "H": 12,
        "I": 16,
        "J": 12,
        "K": 14,
        "L": 40,
    }

    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    numeric_cols_2 = ["F", "G", "I", "K"]
    numeric_cols_4 = ["H", "J"]
    wrap_cols = ["A", "B", "C", "D", "E", "L"]

    styled_last_row = ws.max_row
    for row in range(first_data_row, styled_last_row + 1):
        for col in wrap_cols:
            ws[f"{col}{row}"].alignment = wrap_alignment

        for col in numeric_cols_2:
            ws[f"{col}{row}"].number_format = "0.00"

        for col in numeric_cols_4:
            ws[f"{col}{row}"].number_format = "0.0000"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = (
        f"A1:{get_column_letter(ws.max_column)}"
        f"{last_data_row if has_data_rows else 1}"
    )

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{_safe_filename(report.title)}.xlsx"
    return output, filename